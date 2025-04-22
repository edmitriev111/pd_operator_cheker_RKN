import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import threading
import pandas as pd
from pd_checker import setup_driver, check_operator_status
import os
import sys
import datetime
from openpyxl.styles import PatternFill

class PDCheckerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("PD Checker")
        self.root.geometry("800x600")
        
        # Определение тем
        self.themes = {
            "dark": {
                "bg": "#1e1e1e",
                "fg": "white",
                "button_bg": "#404040",
                "button_fg": "white",
                "log_bg": "#252526",
                "log_fg": "white"
            },
            "light": {
                "bg": "white",
                "fg": "black",
                "button_bg": "#f0f0f0",
                "button_fg": "black",
                "log_bg": "white",
                "log_fg": "black"
            }
        }
        
        self.current_theme = "dark"  # Тема по умолчанию
        self.driver = None
        self.results = []
        
        sys.stdout = self
        sys.stderr = self
        
        self.checker = None
        self.check_thread = None
        self.is_running = False
        
        self.create_widgets()
        self.apply_theme()
        
    def create_widgets(self):
        # Верхняя панель с кнопками и темой
        top_frame = ttk.Frame(self.root, padding="5")
        top_frame.pack(fill=tk.X)
        
        # Тема и информация о разработчике
        theme_frame = ttk.Frame(top_frame)
        theme_frame.pack(side=tk.RIGHT, padx=5)
        
        ttk.Label(theme_frame, text="Тема:").pack(side=tk.LEFT)
        self.theme_var = tk.StringVar(value="dark")
        theme_menu = ttk.OptionMenu(theme_frame, self.theme_var, "dark", 
                                  "dark", "light", 
                                  command=self.change_theme)
        theme_menu.pack(side=tk.LEFT, padx=5)
        
        # Кнопки
        button_frame = ttk.Frame(self.root, padding="5")
        button_frame.pack(fill=tk.X)
        
        # Заменяем ttk.Button на tk.Button
        self.select_btn = tk.Button(button_frame, text="Выбрать файл",
                                   command=self.select_file,
                                   relief="raised",
                                   bd=1)
        self.select_btn.pack(side=tk.LEFT, padx=5)
        
        self.start_btn = tk.Button(button_frame, text="Запустить проверку",
                                  command=self.start_check,
                                  state=tk.DISABLED,
                                  relief="raised",
                                  bd=1)
        self.start_btn.pack(side=tk.LEFT, padx=5)
        
        self.stop_btn = tk.Button(button_frame, text="Стоп",
                                 command=self.stop_check,
                                 state=tk.DISABLED,
                                 relief="raised",
                                 bd=1)
        self.stop_btn.pack(side=tk.LEFT, padx=5)
        
        self.save_btn = tk.Button(button_frame, text="Сохранить результат",
                                 command=self.save_results,
                                 state=tk.DISABLED,
                                 relief="raised",
                                 bd=1)
        self.save_btn.pack(side=tk.LEFT, padx=5)
        
        # Добавляем новые кнопки
        self.stats_btn = tk.Button(button_frame, text="Статистика",
                              command=self.show_statistics,
                              relief="raised",
                              bd=1)
        self.stats_btn.pack(side=tk.LEFT, padx=5)
    
        self.search_btn = tk.Button(button_frame, text="Поиск",
                               command=self.search_results,
                               relief="raised",
                               bd=1)
        self.search_btn.pack(side=tk.LEFT, padx=5)
    
        # Меню сохранения
        save_menu = tk.Menu(button_frame, tearoff=0)
        save_menu.add_command(label="Excel", 
                         command=lambda: self.save_results_as('excel'))
        save_menu.add_command(label="CSV", 
                         command=lambda: self.save_results_as('csv'))
        save_menu.add_command(label="JSON", 
                         command=lambda: self.save_results_as('json'))
    
        self.save_btn.configure(command=lambda: save_menu.post(
            self.save_btn.winfo_rootx(),
            self.save_btn.winfo_rooty() + self.save_btn.winfo_height()
        ))
        
        # Область лога
        log_frame = ttk.Frame(self.root)
        log_frame.pack(fill=tk.BOTH, expand=True, padx=5, pady=5)
        
        self.scrollbar = ttk.Scrollbar(log_frame)
        self.scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.log_text = tk.Text(log_frame, height=30, width=90,
                               wrap=tk.WORD,
                               yscrollcommand=self.scrollbar.set,
                               font=('Consolas', 10))
        self.log_text.pack(fill=tk.BOTH, expand=True)
        
        self.scrollbar.config(command=self.log_text.yview)
        
        # Прогресс бар
        self.progress = ttk.Progressbar(self.root, mode='determinate')
        self.progress.pack(fill=tk.X, padx=5, pady=5)
        
        # Статус бар и информация о разработчике
        status_frame = ttk.Frame(self.root)
        status_frame.pack(fill=tk.X, padx=5, pady=2)
        
        self.status_text = tk.Text(status_frame, 
                              height=1,
                              relief=tk.SUNKEN,
                              font=('Consolas', 9))
        self.status_text.pack(side=tk.LEFT, fill=tk.X, expand=True)
    
        # Делаем поле только для чтения, но с возможностью копирования
        self.status_text.configure(state='disabled')
        
        # Информация о разработчике
        dev_label = ttk.Label(status_frame, 
                             text="Разработчик: e.dmitriev.self@outlook.com",
                             padding=(5, 2))
        dev_label.pack(side=tk.RIGHT)

    def write(self, text):
        """Метод для перенаправления stdout/stderr в лог"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {text}")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def flush(self):
        """Требуется для перенаправления stdout"""
        pass

    def change_theme(self, theme_name):
        self.current_theme = theme_name
        self.apply_theme()
        
    def apply_theme(self):
        theme = self.themes[self.current_theme]
        
        self.root.configure(bg=theme["bg"])
        self.log_text.configure(bg=theme["log_bg"], fg=theme["log_fg"])
        
        # Настройка цветов для кнопок
        buttons = [self.select_btn, self.start_btn, self.stop_btn, self.save_btn]
        for button in buttons:
            button.configure(
                bg=theme["button_bg"],
                fg=theme["button_fg"],
                activebackground="#404040" if self.current_theme == "dark" else "#e0e0e0",
                activeforeground=theme["button_fg"]
            )
        
        # Обновляем стили для ttk виджетов
        style = ttk.Style()
        style.configure("TFrame", background=theme["bg"])
        style.configure("TLabel", background=theme["bg"], foreground=theme["fg"])
        style.configure("TMenubutton", 
                       background=theme["button_bg"],
                       foreground=theme["button_fg"])
        
        self.status_text.configure(
            bg=theme["log_bg"],
            fg=theme["log_fg"],
            insertbackground=theme["fg"]
        )

    def select_file(self):
        self.input_file = filedialog.askopenfilename(
            filetypes=[("Text files", "*.txt")],
            title="Выберите файл с ИНН"
        )
        if self.input_file:
            self.log_message(f"Выбран файл: {self.input_file}", "SUCCESS")
            self.start_btn.config(state=tk.NORMAL)
            self.update_status(f"Готов к проверке: {os.path.basename(self.input_file)}")

    def start_check(self):
        if not hasattr(self, 'input_file'):
            messagebox.showerror("Ошибка", "Сначала выберите файл!")
            return
            
        self.is_running = True
        self.results = []
        
        # Добавляем проверку успешности инициализации драйвера
        self.driver = setup_driver()
        if self.driver is None:
            messagebox.showerror("Ошибка", 
                "Не удалось инициализировать ChromeDriver.\n" +
                "Проверьте:\n" +
                "1. Установлен ли Google Chrome\n" +
                "2. Обновите Chrome до последней версии\n" +
                "3. Попробуйте запустить от имени администратора")
            return
        
        self.start_btn.config(state=tk.DISABLED)
        self.stop_btn.config(state=tk.NORMAL)
        self.save_btn.config(state=tk.DISABLED)
        
        self.check_thread = threading.Thread(target=self.run_check)
        self.check_thread.start()

    def run_check(self):
        try:
            self.clear_log()
            self.log_message("Начало проверки...", "INFO")
            self.update_status("Выполняется проверка...")
            
            with open(self.input_file, 'r') as f:
                inn_list = [line.strip() for line in f if line.strip()]
                
            total = len(inn_list)
            for i, inn in enumerate(inn_list):
                if not self.is_running:
                    self.log_message("Проверка остановлена пользователем", "WARNING")
                    break
                    
                self.progress['value'] = (i + 1) / total * 100
                self.log_message(f"Проверка ИНН {inn} ({i+1}/{total})", "INFO")
                
                try:
                    result = check_operator_status(self.driver, inn)
                    self.results.append({'inn': inn, 'data': result})
                    if result:
                        self.log_message(f"Найдено для ИНН {inn}", "SUCCESS")
                    else:
                        self.log_message(f"ИНН {inn} не найден в реестре", "WARNING")
                except Exception as e:
                    self.log_message(f"Ошибка при проверке ИНН {inn}: {str(e)}", "ERROR")
                    
            self.log_message("Проверка завершена", "SUCCESS")
            self.update_status("Проверка завершена")
            
        except Exception as e:
            self.log_message(f"Критическая ошибка: {str(e)}", "ERROR")
            self.update_status("Произошла ошибка")
        finally:
            if self.driver:
                self.driver.quit()
            self.stop_btn.config(state=tk.DISABLED)
            self.start_btn.config(state=tk.NORMAL)
            self.save_btn.config(state=tk.NORMAL)

    def stop_check(self):
        self.is_running = False
        if self.driver:
            self.driver.quit()
        self.log_message("Остановка проверки...", "WARNING")
        self.stop_btn.config(state=tk.DISABLED)
        self.save_btn.config(state=tk.NORMAL)

    def save_results(self):
        if not self.results:
            messagebox.showerror("Ошибка", "Нет данных для сохранения!")
            return
            
        save_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx")],
            title="Сохранить результаты"
        )
        
        if save_path:
            try:
                df = self.create_dataframe()
                self.save_excel_report(df, save_path)
                self.log_message(f"Результаты сохранены в: {save_path}", "SUCCESS")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении: {str(e)}")

    def save_results_as(self, format_type):
        if not self.results:
            messagebox.showerror("Ошибка", "Нет данных для сохранения!")
            return
            
        formats = {
            'excel': ('.xlsx', 'Excel files'),
            'csv': ('.csv', 'CSV files'),
            'json': ('.json', 'JSON files')
        }
        
        ext, desc = formats[format_type]
        save_path = filedialog.asksaveasfilename(
            defaultextension=ext,
            filetypes=[(desc, f"*{ext}")],
            title="Сохранить результаты"
        )
        
        if save_path:
            df = self.create_dataframe()
            try:
                if format_type == 'excel':
                    self.save_excel_report(df, save_path)
                elif format_type == 'csv':
                    df.to_csv(save_path, index=False, encoding='utf-8-sig')
                elif format_type == 'json':
                    df.to_json(save_path, orient='records', force_ascii=False, indent=2)
                self.log_message(f"Результаты сохранены в: {save_path}", "SUCCESS")
            except Exception as e:
                messagebox.showerror("Ошибка", f"Ошибка при сохранении: {str(e)}")

    def create_dataframe(self):
        data = []
        for result in self.results:
            inn = result['inn']  # Получаем ИНН из результата
            status = "Найден" if result['data'] else "Не найден"  # Определяем статус
            
            if result['data']:
                name_inn = result['data']['name_inn']
                name = name_inn.split('ИНН:')[0].strip()
                
                data.append({
                    'ИНН': inn,
                    'Статус': status,
                    'Наименование': name,
                    'Тип оператора': result['data']['operator_type'],
                    'Основание включения': result['data']['inclusion_basis'],
                    'Дата регистрации': result['data']['registration_date'],
                    'Дата начала обработки': result['data']['processing_start_date'],
                    'Ответственный за ПД': result['data'].get('responsible_person', 'Не указан'),
                    'Email': result['data'].get('email', ''),
                    'Ссылка на карточку': result['data'].get('url', '')
                })
            else:
                # Добавляем запись для ненайденного ИНН
                data.append({
                    'ИНН': inn,
                    'Статус': status,
                    'Наименование': '-',
                    'Тип оператора': '-',
                    'Основание включения': '-',
                    'Дата регистрации': '-',
                    'Дата начала обработки': '-',
                    'Ответственный за ПД': '-',
                    'Email': '-',
                    'Ссылка на карточку': '-'
                })
        
        # Обновляем порядок столбцов
        columns = [
            'ИНН',
            'Статус',
            'Наименование',
            'Тип оператора',
            'Основание включения',
            'Дата регистрации',
            'Дата начала обработки',
            'Ответственный за ПД',
            'Email',
            'Ссылка на карточку'
        ]
        
        return pd.DataFrame(data, columns=columns)

    def save_excel_report(self, df, filename):
        with pd.ExcelWriter(filename, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Операторы')
            
            worksheet = writer.sheets['Операторы']
            
            # Добавляем условное форматирование для столбца "Статус"
            status_column = 'B'  # Столбец B для "Статус"
            for row in range(2, len(df) + 2):  # +2 так как Excel начинается с 1 и есть заголовок
                cell = f"{status_column}{row}"
                if worksheet[cell].value == "Найден":
                    worksheet[cell].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")
                else:
                    worksheet[cell].fill = PatternFill(start_color="FFB6C1", end_color="FFB6C1", fill_type="solid")
            
            # Настраиваем ширину столбцов
            for column in worksheet.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                worksheet.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # Добавляем фильтр
            worksheet.auto_filter.ref = worksheet.dimensions

    def log_message(self, message, level="INFO"):
        """Логирование с уровнями и цветами"""
        timestamp = datetime.datetime.now().strftime("%H:%M:%S")
        
        colors = {
            "INFO": "white",
            "ERROR": "red",
            "SUCCESS": "green",
            "WARNING": "yellow"
        }
        
        self.log_text.tag_config(level, foreground=colors[level])
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n", level)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def update_status(self, message):
        """Обновление статус бара с возможностью копирования"""
        self.status_text.configure(state='normal')
        self.status_text.delete(1.0, tk.END)
        self.status_text.insert(tk.END, message)
        self.status_text.configure(state='disabled')
        
    def clear_log(self):
        """Очистка лога"""
        self.log_text.delete(1.0, tk.END)

    def search_results(self):
        if not self.results:
            messagebox.showinfo("Информация", "Нет данных для поиска!")
            return
            
        search_window = tk.Toplevel(self.root)
        search_window.title("Поиск по результатам")
        search_window.geometry("400x300")
        
        ttk.Label(search_window, text="Введите текст для поиска:").pack(pady=5)
        search_entry = ttk.Entry(search_window, width=40)
        search_entry.pack(pady=5)
        
        results_text = tk.Text(search_window, height=15, width=45)
        results_text.pack(pady=5)
        
        def perform_search():
            query = search_entry.get().lower()
            results_text.delete(1.0, tk.END)
            
            found = False
            df = self.create_dataframe()
            for _, row in df.iterrows():
                if any(query in str(value).lower() for value in row):
                    results_text.insert(tk.END, f"ИНН: {row['ИНН']}\n")
                    results_text.insert(tk.END, f"Наименование: {row['Наименование']}\n")
                    results_text.insert(tk.END, "-" * 40 + "\n")
                    found = True
                    
            if not found:
                results_text.insert(tk.END, "Ничего не найдено")
        
        ttk.Button(search_window, text="Найти", command=perform_search).pack(pady=5)

    def show_statistics(self):
        if not self.results:
            messagebox.showinfo("Информация", "Нет данных для анализа!")
            return
            
        df = self.create_dataframe()
        
        stats_window = tk.Toplevel(self.root)
        stats_window.title("Статистика")
        stats_window.geometry("400x300")
        
        stats_text = tk.Text(stats_window, height=15, width=45)
        stats_text.pack(pady=5)
        
        total = len(df)
        found = len(df[df['Статус'] == 'Найден'])
        not_found = len(df[df['Статус'] == 'Не найден'])
        
        stats = (
            f"Общая статистика:\n"
            f"{'='*40}\n"
            f"Всего проверено: {total}\n"
            f"Найдено: {found} ({found/total*100:.1f}%)\n"
            f"Не найдено: {not_found} ({not_found/total*100:.1f}%)\n"
            f"\nТоп операторов по типу:\n"
            f"{'='*40}\n"
        )
        
        if found > 0:
            type_stats = df[df['Статус'] == 'Найден']['Тип оператора'].value_counts()
            for type_name, count in type_stats.items():
                stats += f"{type_name}: {count}\n"
        
        stats_text.insert(tk.END, stats)
        stats_text.configure(state='disabled')

if __name__ == "__main__":
    root = tk.Tk()
    app = PDCheckerGUI(root)
    root.mainloop()